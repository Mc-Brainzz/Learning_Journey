import requests
import openpyxl

# Hugging Face API details
 # Replace with your API token

# Function to call Hugging Face API
def generate_email(professor_name, research_paper, research_field):
    prompt = (
        f"Write a professional and personalized email to Professor {professor_name}. "
        f"Discuss their research paper titled '{research_paper}' in the field of {research_field}. "
        f"Make it polite, appreciative, and professional."
    )
    response = requests.post(
        API_URL,
        headers=headers,
        json={"inputs": prompt}
    )

    if response.status_code == 200:
        return response.json()["generated_text"]
    else:
        print("Error:", response.status_code, response.text)
        return None

# Function to update the Excel sheet with email details
def update_excel(file_path, professor_name, research_paper, research_field, email_body):
    # Load or create the workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the first sheet or create one
    if "Sheet1" in workbook.sheetnames:
        sheet = workbook["Sheet1"]
    else:
        sheet = workbook.create_sheet("Sheet1")

    # Add headers if the sheet is empty
    if sheet.max_row == 1:
        sheet.append(["Professor Name", "Research Paper", "Research Field", "Email Body"])

    # Append the data
    sheet.append([professor_name, research_paper, research_field, email_body])

    # Save the workbook
    workbook.save(file_path)

# Main function to automate the workflow
def main():
    # Input details
    professor_name = input("Enter Professor's Name: ")
    research_paper = input("Enter Research Paper Title: ")
    research_field = input("Enter Research Field: ")

    # Generate the email
    email_body = generate_email(professor_name, research_paper, research_field)

    if email_body:
        print("\nGenerated Email:\n")
        print(email_body)

        # Update the Excel sheet
        excel_file = "emails_sent.xlsx"
        update_excel(excel_file, professor_name, research_paper, research_field, email_body)
        print(f"\nEmail details saved to {excel_file}")

# Run the program
if __name__ == "__main__":
    main()
